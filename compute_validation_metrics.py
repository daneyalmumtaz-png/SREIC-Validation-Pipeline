"""
Stage 2: load the Excel from Stage 1, run Model.py on each event,
compute Pearson r, MAE, MAPE, mean residual — each with bootstrap CIs
— plus a sign test for systematic bias. Append the results back into
the Excel and write a markdown report.

Run this after fetch_validation_data.py.
"""

from __future__ import annotations

import logging
import math
import random
import sys
from pathlib import Path

import config

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
)
log = logging.getLogger("sreic.metrics")


def _import_user_model():
    """Import Model.py from cwd or this script's folder. Edit run_model
    below if your Model.py uses different function names."""
    sys.path.insert(0, str(Path.cwd()))
    sys.path.insert(0, str(Path(__file__).parent))
    try:
        import Model as M
        return M
    except ImportError:
        pass
    try:
        from Model import compute_impact, load_acceleration_table
        import types
        m = types.SimpleNamespace()
        m.compute_impact = compute_impact
        m.load_acceleration_table = load_acceleration_table
        return m
    except ImportError as e:
        raise SystemExit(f"Cannot import Model.py: {e}")


def run_model(M, v0_kmh, vt_kmh, vr_kmh, L_m, accel_table):
    """Wrapper around Model.compute_impact. Returns (traversal, baseline, delay)."""
    if hasattr(M, "compute_impact"):
        out = M.compute_impact(
            v0=v0_kmh, vt=vt_kmh, vr=vr_kmh, L=L_m,
            decel=config.SM3_DECEL_MS2,
            accel_table=accel_table,
            dt=config.KINEMATIC_DT,
        )
        if isinstance(out, dict):
            tr = out.get("traversal_s") or out.get("t_with") or out["restricted_time_s"]
            bs = out.get("baseline_s") or out.get("t_without") or out["baseline_time_s"]
            dl = out.get("delay_s") or (tr - bs)
            return float(tr), float(bs), float(dl)
        if isinstance(out, (tuple, list)) and len(out) >= 3:
            return float(out[0]), float(out[1]), float(out[2])
    raise RuntimeError("compute_impact() returned an unexpected shape — "
                       "edit run_model() in compute_validation_metrics.py.")


# --- statistics ---

def pearson_r(xs, ys):
    n = len(xs)
    if n < 2:
        return float("nan")
    mx = sum(xs) / n
    my = sum(ys) / n
    sxy = sum((x - mx) * (y - my) for x, y in zip(xs, ys))
    sxx = sum((x - mx) ** 2 for x in xs)
    syy = sum((y - my) ** 2 for y in ys)
    den = math.sqrt(sxx * syy)
    return sxy / den if den > 0 else float("nan")


def bootstrap_ci(values_xy, stat_fn, B=10000, alpha=0.05, seed=1234):
    """Paired non-parametric bootstrap. Returns (point, lo, hi)."""
    rng = random.Random(seed)
    n = len(values_xy)
    if n < 2:
        return (float("nan"), float("nan"), float("nan"))
    xs = [p[0] for p in values_xy]
    ys = [p[1] for p in values_xy]
    point = stat_fn(xs, ys)
    samples = []
    for _ in range(B):
        idx = [rng.randrange(n) for _ in range(n)]
        s = stat_fn([xs[i] for i in idx], [ys[i] for i in idx])
        if not (math.isnan(s) or math.isinf(s)):
            samples.append(s)
    samples.sort()
    lo = samples[int(B * alpha / 2)]
    hi = samples[int(B * (1 - alpha / 2)) - 1]
    return point, lo, hi


def sign_test_p_value(residuals):
    """Two-sided exact binomial sign test under H0 of zero median bias."""
    n = len(residuals)
    pos = sum(1 for r in residuals if r > 0)
    if n == 0:
        return float("nan")
    k = min(pos, n - pos)
    from math import comb
    tail = sum(comb(n, i) for i in range(k + 1)) * 2 / (2 ** n)
    return min(tail, 1.0)


# --- main ---

def main():
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise SystemExit("openpyxl missing — pip install openpyxl")

    M = _import_user_model()

    # acceleration table
    if hasattr(M, "load_acceleration_table"):
        accel = M.load_acceleration_table(config.SM3_ACCEL_TABLE_PATH)
    else:
        accel = _load_accel_csv(config.SM3_ACCEL_TABLE_PATH)
    log.info(f"Loaded acceleration table with {len(accel)} entries")

    wb = load_workbook(config.OUTPUT_EXCEL_PATH)
    ws = wb["validation_events"]
    headers = [c.value for c in ws[1]]
    col = {h: i + 1 for i, h in enumerate(headers)}

    pairs = []
    residuals = []
    abs_pcts = []
    n_processed = 0

    for r_i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        try:
            v0 = float(row[col["v0_kmh"] - 1].value)
            vt = float(row[col["vt_kmh"] - 1].value)
            vr = float(row[col["vr_kmh"] - 1].value)
            L = float(row[col["L_m"] - 1].value)
            obs_traversal = float(row[col["observed_traversal_s"] - 1].value)
        except (TypeError, ValueError):
            log.warning(f"row {r_i} missing input — skipping")
            continue

        try:
            mt, mb, md = run_model(M, v0, vt, vr, L, accel)
        except Exception as e:
            log.warning(f"row {r_i} model call failed: {e}")
            continue

        # observed delay = observed_traversal − model_baseline
        # (no real counterfactual baseline available; the modelled
        # baseline is the only consistent reference)
        obs_delay = obs_traversal - mb
        resid = obs_delay - md
        apct = 100.0 * abs(resid) / abs(obs_delay) if obs_delay != 0 else float("nan")

        ws.cell(row=r_i, column=col["model_traversal_s"], value=round(mt, 2))
        ws.cell(row=r_i, column=col["model_baseline_s"], value=round(mb, 2))
        ws.cell(row=r_i, column=col["model_predicted_delay_s"], value=round(md, 2))
        ws.cell(row=r_i, column=col["observed_delay_s"], value=round(obs_delay, 2))
        ws.cell(row=r_i, column=col["residual_s"], value=round(resid, 2))
        ws.cell(row=r_i, column=col["abs_pct_error"],
                value=round(apct, 2) if not math.isnan(apct) else "")

        pairs.append((md, obs_delay))
        residuals.append(resid)
        if not math.isnan(apct):
            abs_pcts.append(apct)
        n_processed += 1

    log.info(f"Processed {n_processed} events through Model.py")

    summary = _compute_summary(pairs, residuals, abs_pcts)

    if "summary" in wb.sheetnames:
        del wb["summary"]
    ws3 = wb.create_sheet("summary")
    _write_summary_sheet(ws3, summary, Font, PatternFill, Alignment)

    wb.save(config.OUTPUT_EXCEL_PATH)
    log.info(f"Updated {config.OUTPUT_EXCEL_PATH}")

    _write_markdown_report(summary)


def _load_accel_csv(path):
    out = []
    with open(path, "r", encoding="utf-8") as f:
        for ln in f:
            ln = ln.strip()
            if not ln or ln.startswith("#") or ln.lower().startswith("speed"):
                continue
            parts = [p.strip() for p in ln.replace(";", ",").split(",")]
            try:
                out.append((float(parts[0]), float(parts[1])))
            except (ValueError, IndexError):
                continue
    out.sort()
    return out


def _compute_summary(pairs, residuals, abs_pcts):
    n = len(pairs)
    if n < 2:
        return {"n": n, "warning": "insufficient sample size for statistics"}

    r_pt, r_lo, r_hi = bootstrap_ci(pairs, pearson_r)

    bias_pt = sum(residuals) / n
    bias_pairs = [(0.0, r) for r in residuals]
    _, bias_lo, bias_hi = bootstrap_ci(bias_pairs, lambda xs, ys: sum(ys) / len(ys))

    mae_pt = sum(abs(r) for r in residuals) / n
    mae_pairs = [(0.0, abs(r)) for r in residuals]
    _, mae_lo, mae_hi = bootstrap_ci(mae_pairs, lambda xs, ys: sum(ys) / len(ys))

    if abs_pcts:
        mape_pt = sum(abs_pcts) / len(abs_pcts)
        mape_pairs = [(0.0, p) for p in abs_pcts]
        _, mape_lo, mape_hi = bootstrap_ci(mape_pairs, lambda xs, ys: sum(ys) / len(ys))
    else:
        mape_pt = mape_lo = mape_hi = float("nan")

    pos = sum(1 for r in residuals if r > 0)
    neg = sum(1 for r in residuals if r < 0)
    sign_p = sign_test_p_value(residuals)

    return {
        "n": n,
        "pearson_r": {"point": r_pt, "ci_low": r_lo, "ci_high": r_hi},
        "mean_residual_s": {"point": bias_pt, "ci_low": bias_lo, "ci_high": bias_hi},
        "mae_s": {"point": mae_pt, "ci_low": mae_lo, "ci_high": mae_hi},
        "mape_pct": {"point": mape_pt, "ci_low": mape_lo, "ci_high": mape_hi},
        "n_positive_residuals": pos,
        "n_negative_residuals": neg,
        "n_zero_residuals": n - pos - neg,
        "sign_test_p_value": sign_p,
    }


def _fmt(s, fmt=".3f"):
    if not isinstance(s, dict):
        return str(s)
    return f"{s['point']:{fmt}} (95% CI [{s['ci_low']:{fmt}}, {s['ci_high']:{fmt}}])"


def _write_summary_sheet(ws, s, Font, PatternFill, Alignment):
    rows = [
        ["SREIC validation summary"],
        [""],
        ["n", s.get("n")],
        [""],
        ["Pearson r", _fmt(s.get("pearson_r", {}))],
        ["Mean residual (s)", _fmt(s.get("mean_residual_s", {}), ".1f")],
        ["MAE (s)", _fmt(s.get("mae_s", {}), ".1f")],
        ["MAPE (%)", _fmt(s.get("mape_pct", {}), ".1f")],
        [""],
        ["Positive residuals", s.get("n_positive_residuals")],
        ["Negative residuals", s.get("n_negative_residuals")],
        ["Zero residuals", s.get("n_zero_residuals")],
        ["Sign-test p-value", round(s.get("sign_test_p_value", 0), 4)],
        [""],
        ["Bootstrap: paired non-parametric, B=10000, alpha=0.05"],
    ]
    for r in rows:
        ws.append(r)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 50
    ws["A1"].font = Font(bold=True, size=14)


def _write_markdown_report(s):
    n = s.get("n", 0)
    if n < 2:
        return
    p = s["pearson_r"]
    b = s["mean_residual_s"]
    m = s["mae_s"]
    mp = s["mape_pct"]
    md = [
        "# SREIC validation report",
        "",
        "## Sample",
        f"- n = {n} events",
        f"- Date window: {config.DATE_RANGE_START} to {config.DATE_RANGE_END}",
        "",
        "## Headline statistics",
        f"- Pearson r = {_fmt(p)}",
        f"- Mean residual (bias) = {_fmt(b, '.1f')} s",
        f"- MAE = {_fmt(m, '.1f')} s",
        f"- MAPE = {_fmt(mp, '.1f')} %",
        "",
        "## Sign of residuals",
        f"- Positive: {s['n_positive_residuals']}",
        f"- Negative: {s['n_negative_residuals']}",
        f"- Zero: {s['n_zero_residuals']}",
        f"- Two-sided sign-test p = {s['sign_test_p_value']:.4f}",
        "",
        "## Suggested paragraph for Section 4.1",
        (f"Across {n} Pendolino Sm3 telemetry events drawn from the Digitraffic "
         f"open-data feed and matched to independent restriction notifications "
         f"across the period {config.DATE_RANGE_START}–{config.DATE_RANGE_END}, "
         f"predicted and observed restricted-zone delays were positively correlated "
         f"(Pearson r = {p['point']:.3f}, 95% non-parametric bootstrap CI "
         f"[{p['ci_low']:.3f}, {p['ci_high']:.3f}], B = 10 000). "
         f"Mean residual was {b['point']:.1f} s "
         f"(95% CI [{b['ci_low']:.1f}, {b['ci_high']:.1f}]); "
         f"MAE = {m['point']:.1f} s "
         f"(95% CI [{m['ci_low']:.1f}, {m['ci_high']:.1f}]); "
         f"MAPE = {mp['point']:.1f}% "
         f"(95% CI [{mp['ci_low']:.1f}%, {mp['ci_high']:.1f}%]). "
         f"Of {n} residuals, {s['n_negative_residuals']} were negative and "
         f"{s['n_positive_residuals']} positive (two-sided sign-test "
         f"p = {s['sign_test_p_value']:.3f})."),
        "",
    ]
    Path("validation_report.md").write_text("\n".join(md), encoding="utf-8")
    log.info("Wrote validation_report.md")


if __name__ == "__main__":
    main()
