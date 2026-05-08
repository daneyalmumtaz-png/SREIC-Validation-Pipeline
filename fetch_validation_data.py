"""
Stage 1: pull Pendolino telemetry from Digitraffic and TSR ground truth
from Jeti (or an offline file), match speed dips to TSRs, write the
result to Excel.

Run this first. Output goes to compute_validation_metrics.py for the
statistics.
"""

from __future__ import annotations

import csv
import io
import json
import logging
import math
import sys
import time
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any

import requests

import config


# Force UTF-8 stdout so the cp1252 console on Windows doesn't crash on
# arrows in log messages. Took an annoying while to track down.
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    handlers=[
        logging.FileHandler(config.OUTPUT_LOG_PATH, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("sreic")


class RateLimiter:
    def __init__(self, rpm):
        self.min_interval = 60.0 / float(rpm)
        self._last = 0.0

    def wait(self):
        elapsed = time.monotonic() - self._last
        if elapsed < self.min_interval:
            time.sleep(self.min_interval - elapsed)
        self._last = time.monotonic()


_rate = RateLimiter(config.RATE_LIMIT_REQUESTS_PER_MIN)


def _hdr():
    return {
        "Digitraffic-User": config.DIGITRAFFIC_USER_HEADER,
        "Accept-Encoding": "gzip",
        "Connection": "close",
    }


def _get_json(url, params=None, retries=3):
    for i in range(retries):
        _rate.wait()
        try:
            r = requests.get(url, headers=_hdr(), params=params, timeout=60)
            if r.status_code == 200:
                return r.json()
            if r.status_code == 429:
                t = 2 ** i
                log.warning(f"429, sleeping {t}s")
                time.sleep(t)
                continue
            r.raise_for_status()
        except requests.RequestException as e:
            log.warning(f"GET {url} attempt {i+1}: {e}")
            time.sleep(1 + i)
    raise RuntimeError(f"GET {url} failed after {retries} attempts")


def _post_gql(query, variables=None, retries=3):
    payload = {"query": query, "variables": variables or {}}
    headers = _hdr() | {"Content-Type": "application/json"}
    for i in range(retries):
        _rate.wait()
        try:
            r = requests.post(config.GRAPHQL_URL, headers=headers,
                              data=json.dumps(payload), timeout=120)
            if r.status_code == 200:
                body = r.json()
                if "errors" in body:
                    log.warning(f"GraphQL errors: {body['errors']}")
                return body
            if r.status_code == 429:
                t = 2 ** i
                log.warning(f"429, sleeping {t}s")
                time.sleep(t)
                continue
            r.raise_for_status()
        except requests.RequestException as e:
            log.warning(f"POST gql attempt {i+1}: {e}")
            time.sleep(1 + i)
    raise RuntimeError(f"GraphQL POST failed after {retries} attempts")


def haversine_m(lat1, lon1, lat2, lon2):
    R = 6371008.8
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    a = math.sin(dp/2)**2 + math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
    return 2 * R * math.asin(math.sqrt(a))


def parse_iso_utc(ts):
    if ts.endswith("Z"):
        ts = ts[:-1] + "+00:00"
    return datetime.fromisoformat(ts).astimezone(timezone.utc)


def extract_lat_lon(g):
    """Pull (lat, lon) from whatever Digitraffic decides to return for a
    location field. Multiple encodings exist in the wild — flat array,
    GeoJSON, {lat,lon}, {x,y}. Trying them in order. Returns
    (None, None) on miss."""
    if g is None:
        return None, None
    if isinstance(g, list) and len(g) >= 2:
        try:
            return float(g[1]), float(g[0])
        except (TypeError, ValueError):
            return None, None
    if isinstance(g, dict):
        c = g.get("coordinates")
        if isinstance(c, list) and len(c) >= 2:
            try:
                return float(c[1]), float(c[0])
            except (TypeError, ValueError):
                pass
        if "longitude" in g and "latitude" in g:
            try:
                return float(g["latitude"]), float(g["longitude"])
            except (TypeError, ValueError):
                pass
        if "x" in g and "y" in g:
            try:
                return float(g["y"]), float(g["x"])
            except (TypeError, ValueError):
                pass
        if "lat" in g and ("lon" in g or "lng" in g):
            try:
                return float(g["lat"]), float(g.get("lon", g.get("lng")))
            except (TypeError, ValueError):
                pass
    return None, None


# --- data classes ---

@dataclass
class TSRRecord:
    notification_id: str
    version: int
    valid_from: datetime
    valid_to: datetime | None
    restricted_speed_kmh: float
    length_m: float
    centroid_lat: float
    centroid_lon: float
    state: str = ""
    raw_geometry: Any | None = None


@dataclass
class TrainTimetableRow:
    station_short_code: str
    station_name: str
    station_lat: float | None
    station_lon: float | None
    scheduled_time: datetime
    actual_time: datetime | None
    is_commercial_stop: bool
    row_type: str


@dataclass
class TrainTrace:
    train_number: str
    departure_date: str
    operator: str
    train_type: str
    timetable: list[TrainTimetableRow] = field(default_factory=list)
    locations: list[dict] = field(default_factory=list)


@dataclass
class ValidationEvent:
    train_number: str
    departure_date: str
    train_type: str
    tsr_id: str
    tsr_version: int
    tsr_state: str
    vr_kmh: float
    L_m: float
    tsr_centroid_lat: float
    tsr_centroid_lon: float
    tsr_valid_from: datetime
    tsr_valid_to: datetime | None
    v0_kmh: float
    vt_kmh: float
    observed_entry_ts: datetime
    observed_exit_ts: datetime
    observed_traversal_s: float
    n_telemetry_samples_in_dip: int
    distance_to_jeti_centroid_m: float


# --- TSR ground truth (online or offline) ---

def fetch_tsrs(start, end):
    if config.OFFLINE_TSR_PATH:
        log.info(f"Loading TSRs from offline file: {config.OFFLINE_TSR_PATH}")
        return load_offline_tsrs(config.OFFLINE_TSR_PATH, start, end)
    return fetch_jeti_tsrs(start, end)


def fetch_jeti_tsrs(start, end):
    log.info(f"Fetching Jeti TSR notifications for {start} -> {end}")
    raw = _get_json(config.JETI_RESTRICTIONS_URL)

    if isinstance(raw, list):
        records = raw
    elif isinstance(raw, dict) and "features" in raw:
        records = raw["features"]
    else:
        records = raw.get("notifications", []) if isinstance(raw, dict) else []

    log.info(f"Received {len(records)} raw Jeti records")

    tsrs = []
    states = {}
    for rec in records:
        try:
            tsr = _parse_jeti(rec)
        except Exception as e:
            log.debug(f"unparseable Jeti record: {e}")
            continue
        if tsr is None:
            continue

        states[tsr.state] = states.get(tsr.state, 0) + 1

        if tsr.state not in config.ALLOWED_TSR_STATES:
            continue
        if tsr.valid_from.date() > end:
            continue
        if tsr.valid_to is not None and tsr.valid_to.date() < start:
            continue
        if tsr.length_m < config.MIN_RESTRICTION_LENGTH_M:
            continue
        if tsr.length_m > config.MAX_RESTRICTION_LENGTH_M:
            continue
        tsrs.append(tsr)

    log.info(f"State distribution in raw data: {states}")
    log.info(f"After all filters: {len(tsrs)} TSRs")
    return tsrs


def _parse_jeti(rec):
    nid = rec.get("id")
    if not nid:
        return None

    state = str(rec.get("state", ""))

    vfrom = rec.get("startDate")
    vto = rec.get("finished") or rec.get("endDate") or rec.get("plannedEnd")
    if not vfrom:
        return None
    try:
        valid_from = parse_iso_utc(vfrom)
        valid_to = parse_iso_utc(vto) if vto else None
    except (ValueError, AttributeError):
        return None

    locs = rec.get("locations") or []
    if not locs:
        return None

    # Walk locations -> identifierRanges -> speedLimit + polyline.
    # Take the first one we find that has both.
    speed = None
    poly = None
    for lb in locs:
        for ir in (lb.get("identifierRanges") or []):
            sl = ir.get("speedLimit")
            if not sl or sl.get("speed") is None:
                continue
            speed = float(sl["speed"])
            raw_poly = ir.get("location")
            if raw_poly and isinstance(raw_poly, list) and raw_poly:
                # sometimes nested one level: [[[lon,lat],...]]
                if isinstance(raw_poly[0], list) and raw_poly[0] \
                        and isinstance(raw_poly[0][0], list):
                    poly = raw_poly[0]
                else:
                    poly = raw_poly
            break
        if speed is not None:
            break

    if speed is None:
        return None

    if poly and len(poly) >= 2:
        L = _polyline_length(poly)
        lat, lon = _polyline_centroid(poly)
    else:
        # fall back to top-level "location" centroid
        flat = rec.get("location")
        if not flat or len(flat) < 2:
            return None
        try:
            lon, lat = float(flat[0]), float(flat[1])
        except (TypeError, ValueError):
            return None
        L = 0.0

    return TSRRecord(
        notification_id=str(nid),
        version=int(rec.get("version", 1)),
        valid_from=valid_from,
        valid_to=valid_to,
        restricted_speed_kmh=speed,
        length_m=L,
        centroid_lat=lat,
        centroid_lon=lon,
        state=state,
        raw_geometry=poly,
    )


def _polyline_length(poly):
    total = 0.0
    for i in range(1, len(poly)):
        try:
            lon1, lat1 = poly[i-1][0], poly[i-1][1]
            lon2, lat2 = poly[i][0], poly[i][1]
            total += haversine_m(lat1, lon1, lat2, lon2)
        except (TypeError, IndexError):
            continue
    return total


def _polyline_centroid(poly):
    n = len(poly)
    if n == 1:
        return float(poly[0][1]), float(poly[0][0])
    return sum(p[1] for p in poly) / n, sum(p[0] for p in poly) / n


def load_offline_tsrs(path, start, end):
    """Load TSRs from a CSV. Expected columns:
    id, version, valid_from, valid_to, restricted_speed_kmh, length_m,
    centroid_lat, centroid_lon, state.

    Adapt this if your RAIDE export uses different column names."""
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Offline TSR file not found: {path}")

    tsrs = []
    with p.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            try:
                vto_raw = row.get("valid_to", "").strip()
                tsr = TSRRecord(
                    notification_id=str(row["id"]),
                    version=int(row.get("version", 1)),
                    valid_from=parse_iso_utc(row["valid_from"]),
                    valid_to=parse_iso_utc(vto_raw) if vto_raw else None,
                    restricted_speed_kmh=float(row["restricted_speed_kmh"]),
                    length_m=float(row["length_m"]),
                    centroid_lat=float(row["centroid_lat"]),
                    centroid_lon=float(row["centroid_lon"]),
                    state=str(row.get("state", "ACTIVE")),
                )
            except (KeyError, ValueError) as e:
                log.debug(f"bad offline row: {e}")
                continue

            if tsr.state not in config.ALLOWED_TSR_STATES:
                continue
            if tsr.valid_from.date() > end:
                continue
            if tsr.valid_to is not None and tsr.valid_to.date() < start:
                continue
            if tsr.length_m < config.MIN_RESTRICTION_LENGTH_M:
                continue
            if tsr.length_m > config.MAX_RESTRICTION_LENGTH_M:
                continue
            tsrs.append(tsr)

    log.info(f"Loaded {len(tsrs)} offline TSRs intersecting window")
    return tsrs


# --- Pendolino traces from Digitraffic GraphQL ---
#
# Two-stage fetch: get train metadata for the day in one query, then
# pull telemetry per-train. The bulk approach (everything in one query)
# made the Digitraffic database time out — see commit history for the
# painful debugging session.

GQL_METADATA = """
query TrainsByDate($d: Date!) {
  trainsByDepartureDate(departureDate: $d) {
    trainNumber
    departureDate
    cancelled
    deleted
    operator { shortCode }
    trainType { name }
    timeTableRows {
      type
      scheduledTime
      actualTime
      commercialStop
      station { shortCode name }
    }
  }
}
"""

GQL_LOCATIONS = """
query TrainLocs($n: Int!, $d: Date!) {
  train(trainNumber: $n, departureDate: $d) {
    trainNumber
    departureDate
    trainLocations(orderBy: { timestamp: ASCENDING }) {
      speed
      timestamp
      location
    }
  }
}
"""


def fetch_pendolino_traces(target_date):
    body = _post_gql(GQL_METADATA, variables={"d": target_date.isoformat()})
    if "data" not in body or not body["data"]:
        log.warning(f"no data block for {target_date}")
        return []
    rows = body["data"].get("trainsByDepartureDate") or []

    # Filter Pendolinos in Python (StringWhere doesn't support `in`).
    candidates = []
    for row in rows:
        if row.get("cancelled") or row.get("deleted"):
            continue
        op = (row.get("operator") or {}).get("shortCode", "")
        if op != config.OPERATOR_CODE:
            continue
        tt = (row.get("trainType") or {}).get("name", "")
        if config.PENDOLINO_TRAIN_TYPES and tt not in config.PENDOLINO_TRAIN_TYPES:
            continue
        if config.TRAIN_NUMBER_WHITELIST \
                and str(row["trainNumber"]) not in config.TRAIN_NUMBER_WHITELIST:
            continue
        candidates.append(row)

    log.info(f"  -> {len(candidates)} Pendolino candidates from {len(rows)} VR trains")

    traces = []
    for row in candidates:
        try:
            tn = int(row["trainNumber"])
        except (TypeError, ValueError):
            continue
        try:
            lb = _post_gql(GQL_LOCATIONS, variables={"n": tn, "d": target_date.isoformat()})
            tobj = (lb.get("data") or {}).get("train")
            if isinstance(tobj, list):
                tobj = tobj[0] if tobj else None
            row["trainLocations"] = (tobj or {}).get("trainLocations") or []
        except Exception as e:
            log.debug(f"telemetry fetch for train {tn} failed: {e}")
            row["trainLocations"] = []

        try:
            traces.append(_parse_train(row))
        except Exception as e:
            log.debug(f"parse train {tn}: {e}")

    log.info(f"  -> {len(traces)} traces with telemetry for {target_date}")
    return traces


def _parse_train(row):
    tt = []
    for ttr in row.get("timeTableRows") or []:
        st = ttr.get("scheduledTime")
        if not st:
            continue
        at = ttr.get("actualTime")
        sobj = ttr.get("station") or {}
        tt.append(TrainTimetableRow(
            station_short_code=sobj.get("shortCode", ""),
            station_name=sobj.get("name", ""),
            station_lat=None,
            station_lon=None,
            scheduled_time=parse_iso_utc(st),
            actual_time=parse_iso_utc(at) if at else None,
            is_commercial_stop=bool(ttr.get("commercialStop")),
            row_type=ttr.get("type", ""),
        ))

    locs = []
    for loc in row.get("trainLocations") or []:
        ts = loc.get("timestamp")
        sp = loc.get("speed")
        if not (ts and sp is not None):
            continue
        lat, lon = extract_lat_lon(loc.get("location"))
        if lat is None or lon is None:
            continue
        locs.append({
            "timestamp": parse_iso_utc(ts),
            "speed_kmh": float(sp),
            "lat": lat,
            "lon": lon,
        })
    locs.sort(key=lambda x: x["timestamp"])

    return TrainTrace(
        train_number=str(row["trainNumber"]),
        departure_date=row["departureDate"],
        operator=row["operator"]["shortCode"],
        train_type=row["trainType"]["name"],
        timetable=tt,
        locations=locs,
    )


# --- station coordinates (one-off lookup) ---

GQL_STATIONS = "query AllStations { stations { shortCode name location } }"


def fetch_station_coordinates():
    body = _post_gql(GQL_STATIONS)
    out = {}
    for s in (body.get("data") or {}).get("stations") or []:
        sc = s.get("shortCode")
        if not sc:
            continue
        lat, lon = extract_lat_lon(s.get("location"))
        if lat is not None and lon is not None:
            out[sc] = (lat, lon)
    log.info(f"Resolved {len(out)} station coordinates")
    return out


# --- the matching engine (this is where the work happens) ---

def find_validation_events(trace, tsrs, station_coords):
    if len(trace.locations) < 10:
        return []

    locs = trace.locations
    n = len(locs)

    # commercial-stop station coords for this trace's route
    stations = [station_coords[r.station_short_code]
                for r in trace.timetable
                if r.is_commercial_stop and r.station_short_code in station_coords]

    # find dip episodes — contiguous runs below DIP_SPEED_THRESHOLD
    episodes = []
    in_dip = False
    start = None
    for i, loc in enumerate(locs):
        if loc["speed_kmh"] < config.DIP_SPEED_THRESHOLD_KMH:
            if not in_dip:
                in_dip = True
                start = i
        else:
            if in_dip:
                episodes.append((start, i - 1))
                in_dip = False
    if in_dip:
        episodes.append((start, n - 1))

    events = []
    for (s, e) in episodes:
        ev = _evaluate_dip(trace, locs, s, e, tsrs, stations)
        if ev is not None:
            events.append(ev)
    return events


def _evaluate_dip(trace, locs, i_start, i_end, tsrs, stations):
    samples = locs[i_start:i_end + 1]

    duration = (samples[-1]["timestamp"] - samples[0]["timestamp"]).total_seconds()
    if duration < config.MIN_DIP_DURATION_S:
        return None

    # dip centroid
    lat_c = sum(s["lat"] for s in samples) / len(samples)
    lon_c = sum(s["lon"] for s in samples) / len(samples)

    # station-dwell filter
    for (slat, slon) in stations:
        if haversine_m(lat_c, lon_c, slat, slon) < config.STATION_DWELL_BUFFER_M:
            return None

    # find the closest TSR active during the dip's time window
    t0 = samples[0]["timestamp"]
    t1 = samples[-1]["timestamp"]
    tol = timedelta(minutes=config.TEMPORAL_MATCH_TOLERANCE_MIN)
    best = None
    best_d = float("inf")
    for tsr in tsrs:
        if tsr.valid_from > t1 + tol:
            continue
        if tsr.valid_to is not None and tsr.valid_to < t0 - tol:
            continue
        d = haversine_m(lat_c, lon_c, tsr.centroid_lat, tsr.centroid_lon)
        if d < config.SPATIAL_MATCH_TOLERANCE_M and d < best_d:
            best = tsr
            best_d = d
    if best is None:
        return None

    # entry / exit speeds bracketing the dip
    if i_start > 0:
        v0 = max(s["speed_kmh"] for s in locs[max(0, i_start - 5):i_start + 1])
    else:
        v0 = samples[0]["speed_kmh"]
    vt = max(s["speed_kmh"] for s in locs[i_end:min(len(locs) - 1, i_end + 10) + 1])

    if v0 - best.restricted_speed_kmh < config.MIN_SPEED_REDUCTION_KMH:
        return None

    return ValidationEvent(
        train_number=trace.train_number,
        departure_date=trace.departure_date,
        train_type=trace.train_type,
        tsr_id=best.notification_id,
        tsr_version=best.version,
        tsr_state=best.state,
        vr_kmh=best.restricted_speed_kmh,
        L_m=best.length_m,
        tsr_centroid_lat=best.centroid_lat,
        tsr_centroid_lon=best.centroid_lon,
        tsr_valid_from=best.valid_from,
        tsr_valid_to=best.valid_to,
        v0_kmh=v0,
        vt_kmh=vt,
        observed_entry_ts=t0,
        observed_exit_ts=t1,
        observed_traversal_s=duration,
        n_telemetry_samples_in_dip=len(samples),
        distance_to_jeti_centroid_m=best_d,
    )


# --- orchestration ---

def run_pipeline():
    diag = {
        "started_at": datetime.now(timezone.utc).isoformat(),
        "date_range": [config.DATE_RANGE_START.isoformat(),
                       config.DATE_RANGE_END.isoformat()],
        "target_n": config.TARGET_SAMPLE_SIZE,
        "tsr_source": "offline" if config.OFFLINE_TSR_PATH else "jeti_api",
        "allowed_tsr_states": list(config.ALLOWED_TSR_STATES),
        "dates_processed": 0,
        "trains_processed": 0,
        "events_accepted": 0,
        "errors": [],
    }

    try:
        tsrs = fetch_tsrs(config.DATE_RANGE_START, config.DATE_RANGE_END)
    except Exception as e:
        log.error(f"TSR fetch failed: {e}")
        diag["errors"].append(f"tsr: {e}")
        tsrs = []

    diag["tsrs_in_window"] = len(tsrs)
    if not tsrs:
        log.error("No TSRs found — pipeline cannot continue.")
        log.error("Likely causes: (1) offline file missing/empty, "
                  "(2) all records filtered out by state or date, "
                  "(3) live snapshot has no in-force records.")
        _write_diag(diag)
        return

    try:
        stations = fetch_station_coordinates()
    except Exception as e:
        log.warning(f"station fetch failed: {e} — dwell filter disabled")
        stations = {}

    if not stations:
        log.warning("Got 0 station coordinates. Station-dwell filtering "
                    "is DISABLED for this run.")

    events = []
    cur = config.DATE_RANGE_START
    while cur <= config.DATE_RANGE_END and len(events) < config.TARGET_SAMPLE_SIZE:
        log.info(f"--- {cur.isoformat()} ---")
        try:
            traces = fetch_pendolino_traces(cur)
        except Exception as e:
            log.warning(f"trace fetch for {cur} failed: {e}")
            diag["errors"].append(f"{cur}: {e}")
            cur += timedelta(days=1)
            continue

        diag["dates_processed"] += 1
        diag["trains_processed"] += len(traces)

        for trace in traces:
            try:
                evs = find_validation_events(trace, tsrs, stations)
                events.extend(evs)
                diag["events_accepted"] += len(evs)
                if config.SAVE_RAW_TRACES and evs:
                    _save_raw_trace(trace)
            except Exception as e:
                log.debug(f"train {trace.train_number} on {cur}: {e}")

            if len(events) >= config.TARGET_SAMPLE_SIZE:
                log.info(f"Reached N={config.TARGET_SAMPLE_SIZE}, stopping.")
                break

        log.info(f"Cumulative events: {len(events)}")
        cur += timedelta(days=1)

    diag["finished_at"] = datetime.now(timezone.utc).isoformat()
    diag["final_n"] = len(events)
    _write_excel(events)
    _write_diag(diag)
    log.info(f"DONE — {len(events)} events written to {config.OUTPUT_EXCEL_PATH}")


def _save_raw_trace(trace):
    Path(config.RAW_TRACES_DIR).mkdir(parents=True, exist_ok=True)
    fn = Path(config.RAW_TRACES_DIR) / f"{trace.departure_date}_{trace.train_number}.json"
    with open(fn, "w", encoding="utf-8") as f:
        json.dump({
            "train_number": trace.train_number,
            "departure_date": trace.departure_date,
            "train_type": trace.train_type,
            "operator": trace.operator,
            "locations": [
                {"ts": l["timestamp"].isoformat(), "v_kmh": l["speed_kmh"],
                 "lat": l["lat"], "lon": l["lon"]}
                for l in trace.locations
            ],
        }, f, indent=2)


def _write_diag(diag):
    with open(config.OUTPUT_DIAGNOSTICS_PATH, "w", encoding="utf-8") as f:
        json.dump(diag, f, indent=2, default=str)


def _write_excel(events):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        log.error("openpyxl missing — pip install openpyxl")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "validation_events"

    cols = [
        "event_id", "train_number", "departure_date", "train_type",
        "tsr_id", "tsr_version", "tsr_state",
        "vr_kmh", "L_m",
        "tsr_centroid_lat", "tsr_centroid_lon",
        "tsr_valid_from", "tsr_valid_to",
        "v0_kmh", "vt_kmh",
        "observed_entry_ts", "observed_exit_ts",
        "observed_traversal_s",
        "n_telemetry_samples_in_dip",
        "distance_to_jeti_centroid_m",
        "model_traversal_s", "model_baseline_s",
        "model_predicted_delay_s", "observed_delay_s",
        "residual_s", "abs_pct_error",
    ]
    ws.append(cols)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E5F")
        c.alignment = Alignment(horizontal="center")

    for i, ev in enumerate(events, start=1):
        ws.append([
            i, ev.train_number, ev.departure_date, ev.train_type,
            ev.tsr_id, ev.tsr_version, ev.tsr_state,
            ev.vr_kmh, ev.L_m,
            ev.tsr_centroid_lat, ev.tsr_centroid_lon,
            ev.tsr_valid_from.isoformat() if ev.tsr_valid_from else "",
            ev.tsr_valid_to.isoformat() if ev.tsr_valid_to else "",
            ev.v0_kmh, ev.vt_kmh,
            ev.observed_entry_ts.isoformat(),
            ev.observed_exit_ts.isoformat(),
            ev.observed_traversal_s,
            ev.n_telemetry_samples_in_dip,
            ev.distance_to_jeti_centroid_m,
            "", "", "", "", "", "",
        ])

    ws2 = wb.create_sheet("README")
    for r in [
        ["SREIC validation dataset"],
        [""],
        ["validation_events: one row per accepted event."],
        ["TSR fields (vr, L, etc) come from Jeti or RAIDE."],
        ["Telemetry fields (v0, vt, observed time) come from Digitraffic."],
        [""],
        ["Run compute_validation_metrics.py next to fill the model columns."],
    ]:
        ws2.append(r)

    wb.save(config.OUTPUT_EXCEL_PATH)
    log.info(f"Wrote {len(events)} events to {config.OUTPUT_EXCEL_PATH}")


if __name__ == "__main__":
    run_pipeline()
